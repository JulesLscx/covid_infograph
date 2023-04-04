# ' -*- coding: utf-8 -*-
import json
import sys
from variables import Keywords
import pymongo as pm
from openpyxl import load_workbook, Workbook
import os
from connection import SingletonMongoConnection as smc


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def str_to_dict(*, string, last__id):
    """Convertit une string en dictionnaire

    Args:
        string (str): string à convertir

    Returns:
        dict: dictionnaire
    """
    string = string.replace("'", '"')
    string = string.replace('None', 'null')
    try:
        dic = json.loads(string)
    except:
        if last__id:
            dic = {'id': last__id, 'issue': 'unable to load interventions column'}
            insert_one_dic_in_mongo(dic=dic, collection='errors')
        dic = {}
    return dic


def test_get_column_names():
    workbook = load_excel_file()
    sheet = workbook[Keywords.WS_CLINICALTRIALS_OBS.value]
    column_names = get_column_names(sheet_object=sheet)
    print(column_names)


def create_dic_by_sheet(sheet_object: object) -> dict:
    column_names = {}
    for cell in sheet_object[1]:
        if cell.value != None:
            column_names[cell.value] = cell.column
    return column_names


def get_column_names(*, sheet_object) -> list[str]:
    """Get column names from the first row of a sheet object.

    Args:
        sheet_object (Workbook.worksheet.worksheet.Worksheet): The sheet object to get the column names from.

    Returns:
        list[str]: A list of column names.
    """
    column_names = []
    for row in sheet_object.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet_object.max_column):
        for cell in row:
            if cell.value == 'id':
                column_names.append('_id')
            else:
                column_names.append(cell.value)
    return column_names


def create_dic_by_sheet(*, sheet_object):
    result = []
    column_names = get_column_names(sheet_object=sheet_object)
    for row in sheet_object.iter_rows(min_row=2, max_row=get_maximum_rows(sheet_object=sheet_object), min_col=1, max_col=sheet_object.max_column):
        dic = {}
        for i, cell in enumerate(row):
            if column_names[i] == 'interventions':
                if cell.value is not None:
                    dic[column_names[i]] = str_to_dict(
                        string=cell.value, last__id=dic.get('_id'))
                else:
                    dic[column_names[i]] = None
            elif column_names[i] == 'concepts' or column_names[i] == 'meshTerms' or column_names[i] == 'openAccess':
                if cell.value is not None:
                    datas = []
                    for data in cell.value.split('•'):
                        datas.append(data.strip())
                    dic[column_names[i]] = datas
                else:
                    dic[column_names[i]] = None
            elif column_names[i] == 'phase':
                tmp = cell.value
                if tmp is None:
                    dic[column_names[i]] = 'N/A'
                elif type(tmp) == int:
                    dic[column_names[i]] = 'Phase ' + str(tmp)
                elif tmp.upper() == 'N/A':
                    dic[column_names[i]] = 'N/A'
                elif 'phase' in tmp.lower():
                    dic[column_names[i]] = tmp.capitalize()
                else:
                    dic[column_names[i]] = 'N/A'
            elif column_names[i] == 'gender':
                if cell.value is None:
                    dic[column_names[i]] = 'N/A'
                elif cell.value.upper() == 'N/A':
                    dic[column_names[i]] = 'N/A'
                elif cell.value.capitalize() not in ('Male', 'Female', 'All'):
                    dic[column_names[i]] = 'N/A'
                else:
                    dic[column_names[i]] = cell.value.capitalize()
            else:
                dic[column_names[i]] = cell.value
        result.append(dic)
    return result


def insert_dic_in_mongo(*, dic_list, collection):
    """insert toutes les dictionnaires d'une liste dans une collection mongoDB chosie


    Args:
        dic_list (list(dict)): liste de dictionnaire correspondant à chaque ligne du fichier excel
        collection (pymongo.collection): Collection mongoDb obligatoirement choisie

    Returns:
        _id str:Le dernière id inséré dans la collection (donc si pas le dernier de la liste c'est qu'il y a eu une erreur)
    """

    for dic in dic_list:
        try:
            collection.insert_one(dic)
        except Exception as e:
            return dic.get('_id')
    return dic_list[-1].get('_id')


def init_collections():
    collections = [Keywords.T_CLINICALTRIALS_OBS.value, Keywords.T_CLINICALTRIALS_RAND.value,
                   Keywords.T_PUBLICATION_OBS.value, Keywords.T_PUBLICATION_RAND.value, 'errors']
    for collection in collections:
        smc.get_db().drop_collection(collection)
        smc.get_db().create_collection(collection)


def load_excel_file(file_name):
    try:
        file_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        file_dir = os.getcwd()

    file_path = os.path.join(
        file_dir, 'excels', file_name)
    try:
        workbook = load_workbook(filename=file_path)
        workbook.encoding = 'utf-8'
    except FileNotFoundError:
        print('Could not find the file')
        workbook = None

    return workbook


def dump_all_excel_in_mongo(file_name):
    sheets = [Keywords.WS_CLINICALTRIALS_OBS.value, Keywords.WS_CLINICALTRIALS_RAND.value,
              Keywords.WS_PUBLICATION_OBS.value, Keywords.WS_PUBLICATION_RAND.value]
    collections = [Keywords.T_CLINICALTRIALS_OBS.value, Keywords.T_CLINICALTRIALS_RAND.value,
                   Keywords.T_PUBLICATION_OBS.value, Keywords.T_PUBLICATION_RAND.value]
    workbook = load_excel_file(file_name)
    for i, sheet in enumerate(sheets):
        dic_list = create_dic_by_sheet(sheet_object=workbook[sheet])
        try:
            smc.get_db()[collections[i]].insert_many(dic_list, ordered=False)
        except Exception as e:
            for dic in dic_list:
                try:
                    smc.get_db()[collections[i]].insert_one(dic)
                except Exception as e:
                    dic['id'] = dic['_id']
                    dic.pop('_id')
                    dic['error'] = str(e)
                    smc.get_db()['errors'].insert_one(dic)


def test_insert_one_dic_in_mongo():
    dic = create_dic_by_sheet(sheet_object=load_excel_file()[
                              Keywords.WS_CLINICALTRIALS_OBS.value])[1:10]


def insert_one_dic_in_mongo(*, dic, collection):
    smc.get_db()[collection].insert_one(dic)


def __main__():
    file_name = sys.argv[1]
    dump_all_excel_in_mongo(file_name)
    # test = create_dic_by_sheet(sheet_object=load_excel_file()[
    #                            Keywords.WS_CLINICALTRIALS_OBS.value])


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: python3 database_add.py <file_name>')
        print('File should be in the covid_infograph/files/excels folder')
        sys.exit(1)
    if not sys.argv[1].endswith('.xlsx'):
        if sys.argv[1] == 'clean':
            init_collections()
            print('Database cleaned')
            exit(0)
        else:
            print('File should be a .xlsx file or a command check the documentation')
    if sys.argv[1] not in os.listdir(os.path.join(os.getcwd(), 'covid_infograph', 'files', 'excels')):
        print('File ' + sys.argv[1] + ' not found in the excels folder')
    __main__()