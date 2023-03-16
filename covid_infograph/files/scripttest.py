# ' -*- coding: utf-8 -*-
import json
from variables import Keywords
import pymongo as pm
from openpyxl import load_workbook, Workbook
import os
from connection import SingletonMongoConnection as smc


# file = pd.read_excel(
#     './others/20200601_IRIT_clinicalTrials+publications.xlsx', engine='openpyxl')


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def str_to_dict(*, string, last__id):
    """Convertit une string en dictionnaire

    Args:
        string (str): string à convertir

    Returns:
        dict: dictionnaire
    """
    string = string.replace("'", '"')
    string = string.replace('None', 'null')
    try:
        dic = json.loads(string)
    except:
        if last__id:
            dic = {'id': last__id, 'issue': 'unable to load interventions column'}
            insert_one_dic_in_mongo(dic=dic, collection='errors')
        dic = {}
    return dic


def test_get_column_names():
    workbook = load_excel_file()
    sheet = workbook[Keywords.WS_CLINICALTRIALS_OBS.value]
    column_names = get_column_names(sheet_object=sheet)
    print(column_names)


def create_dic_by_sheet(sheet_object: object) -> dict:
    column_names = {}
    for cell in sheet_object[1]:
        if cell.value != None:
            column_names[cell.value] = cell.column
    return column_names


def get_column_names(*, sheet_object) -> list[str]:
    """Get column names from the first row of a sheet object.

    Args:
        sheet_object (Workbook.worksheet.worksheet.Worksheet): The sheet object to get the column names from.

    Returns:
        list[str]: A list of column names.
    """
    column_names = []
    for row in sheet_object.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet_object.max_column):
        for cell in row:
            if cell.value == 'id':
                column_names.append('_id')
            else:
                column_names.append(cell.value)
    return column_names


def create_dic_by_sheet(*, sheet_object):
    result = []
    column_names = get_column_names(sheet_object=sheet_object)
    for row in sheet_object.iter_rows(min_row=2, max_row=get_maximum_rows(sheet_object=sheet_object), min_col=1, max_col=sheet_object.max_column):
        dic = {}
        for i, cell in enumerate(row):
            if column_names[i] == 'interventions':
                if cell.value is not None:
                    dic[column_names[i]] = str_to_dict(
                        string=cell.value, last__id=dic.get('_id'))
                else:
                    dic[column_names[i]] = None
            elif column_names[i] == 'concepts' or column_names[i] == 'meshTerms' or column_names[i] == 'openAccess':
                if cell.value is not None:
                    datas = []
                    for data in cell.value.split('•'):
                        datas.append(data.strip())
                    dic[column_names[i]] = datas
                else:
                    dic[column_names[i]] = None
            elif column_names[i] == 'phase':
                tmp = cell.value
                if tmp is None:
                    dic[column_names[i]] = 'N/A'
                elif type(tmp) == int:
                    dic[column_names[i]] = 'Phase ' + str(tmp)
                elif tmp.upper() == 'N/A':
                    dic[column_names[i]] = 'N/A'
                elif 'phase' in tmp.lower():
                    dic[column_names[i]] = tmp.capitalize()
                else:
                    dic[column_names[i]] = 'N/A'
            else:
                dic[column_names[i]] = cell.value
        result.append(dic)
    return result


def insert_dic_in_mongo(*, dic_list, collection):
    """insert toutes les dictionnaires d'une liste dans une collection mongoDB chosie


    Args:
        dic_list (list(dict)): liste de dictionnaire correspondant à chaque ligne du fichier excel
        collection (pymongo.collection): Collection mongoDb obligatoirement choisie

    Returns:
        _id str:Le dernière id inséré dans la collection (donc si pas le dernier de la liste c'est qu'il y a eu une erreur)
    """
    for dic in dic_list:
        try:
            collection.insert_one(dic)
        except Exception as e:
            return dic.get('_id')
    return dic_list[-1].get('_id')


def init_collections():
    collections = [Keywords.T_CLINICALTRIALS_OBS.value, Keywords.T_CLINICALTRIALS_RAND.value,
                   Keywords.T_PUBLICATION_OBS.value, Keywords.T_PUBLICATION_RAND.value, 'errors']
    for collection in collections:
        smc.get_db().drop_collection(collection)
        smc.get_db().create_collection(collection)


def load_excel_file():
    try:
        file_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        file_dir = os.getcwd()

    file_path = os.path.join(
        file_dir, 'excels', '20200601_IRIT_clinicalTrials+publications.xlsx')

    # Copilot ouvre le fichier en utf-8
    try:
        workbook = load_workbook(filename=file_path)
        workbook.encoding = 'utf-8'
    except FileNotFoundError:
        print('Could not find the file')
        workbook = None

    return workbook


def dump_all_excel_in_mongo():
    sheets = [Keywords.WS_CLINICALTRIALS_OBS.value, Keywords.WS_CLINICALTRIALS_RAND.value,
              Keywords.WS_PUBLICATION_OBS.value, Keywords.WS_PUBLICATION_RAND.value]
    collections = [Keywords.T_CLINICALTRIALS_OBS.value, Keywords.T_CLINICALTRIALS_RAND.value,
                   Keywords.T_PUBLICATION_OBS.value, Keywords.T_PUBLICATION_RAND.value]
    workbook = load_excel_file()
    for i, sheet in enumerate(sheets):
        dic_list = create_dic_by_sheet(sheet_object=workbook[sheet])
        smc.get_db()[collections[i]].insert_many(dic_list)


def test_insert_one_dic_in_mongo():
    dic = create_dic_by_sheet(sheet_object=load_excel_file()[
                              Keywords.WS_CLINICALTRIALS_OBS.value])[1:10]
    print(dic)


def insert_one_dic_in_mongo(*, dic, collection):
    smc.get_db()[collection].insert_one(dic)


def __main__():
    init_collections()
    dump_all_excel_in_mongo()
    # test = create_dic_by_sheet(sheet_object=load_excel_file()[
    #                            Keywords.WS_CLINICALTRIALS_OBS.value])


if __name__ == '__main__':
    __main__()

# '''
 # File "C:\Program Files\WindowsApps\PythonSoftwareFoundation.Python.3.10_3.10.2288.0_x64__qbz5n2kfra8p0\lib\encodings\cp1252.py", line 19, in encode
#    return codecs.charmap_encode(input,self.errors,encoding_table)[0]
# UnicodeEncodeError: 'charmap' codec can't encode character '\u2265' in position 1404: character maps to <undefined>
# How to solve this problem?
# answer:
