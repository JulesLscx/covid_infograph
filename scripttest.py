import datetime
import sys
from variables import Keywords
import pymongo as pm
from openpyxl import load_workbook, Workbook
import os
from urllib.parse import quote_plus
from connection import SingletonMongoConnection as smc
import pandas as pd
from io import StringIO


# file = pd.read_excel(
#     './others/20200601_IRIT_clinicalTrials+publications.xlsx', engine='openpyxl')


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def str_to_dict(*, string):
    """Convertit une string en dictionnaire

    Args:
        string (str): string à convertir

    Returns:
        dict: dictionnaire
    """
    buffer = StringIO(string)
    df = pd.read_csv(buffer, sep=',')
    df.to_dict()
# Génère les tests pour la fonction foundlastrow et foundlastcol


# def test_foundlastrow():
#     file_dir = os.path.dirname(os.path.abspath(__file__))
#     file_path = os.path.join(
#         file_dir, 'excels', '20200601_IRIT_clinicalTrials+publications.xlsx')
#     workbook = load_workbook(filename=file_path)
#     sheet = workbook[Keywords.WS_CLINICALTRIALS_OBS.value]
#     last = get_maximum_rows(sheet_object=sheet)
#     print(last, sheet.max_column)

def test_get_column_names():
    workbook = load_excel_file()
    sheet = workbook[Keywords.WS_CLINICALTRIALS_OBS.value]
    column_names = get_column_names(sheet_object=sheet)
    print(column_names)


def test_create_dic_by_sheet():
    workbook = load_excel_file()
    sheet = workbook[Keywords.WS_CLINICALTRIALS_OBS.value]
    column_names = create_dic_by_sheet(sheet_object=sheet)
    print(column_names)


def get_column_names(*, sheet_object):
    column_names = []
    for row in sheet_object.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet_object.max_column):
        for cell in row:
            if cell.value == 'id':
                column_names.append('_id')
            else:
                column_names.append(cell.value)
    return column_names


def create_dic_by_sheet(*, sheet_object):
    result = []
    column_names = get_column_names(sheet_object=sheet_object)
    for row in sheet_object.iter_rows(min_row=1, max_row=get_maximum_rows(sheet_object=sheet_object), min_col=1, max_col=sheet_object.max_column):
        dic = {}
        for i, cell in enumerate(row):
            if column_names[i] == 'interventions':
                if cell.value is not None:
                    dic[column_names[i]] = str_to_dict(
                        string=cell.value)
                else:
                    dic[column_names[i]] = None
            if isinstance(cell.value, str):
                dic[column_names[i]] = cell.value.encode('utf-8')
            else:
                dic[column_names[i]] = cell.value
        result.append(dic)
    return result


def insert_dic_in_mongo(*, dic_list, collection):
    """insert toutes les dictionnaires d'une liste dans une collection mongoDB chosie


    Args:
        dic_list (list(dict)): liste de dictionnaire correspondant à chaque ligne du fichier excel
        collection (pymongo.collection): Collection mongoDb obligatoirement choisie

    Returns:
        _id str:Le dernière id inséré dans la collection (donc si pas le dernier de la liste c'est qu'il y a eu une erreur)
    """
    for dic in dic_list:
        try:
            collection.insert_one(dic)
        except Exception as e:
            return dic.get('_id')
    return dic_list[-1].get('_id')


def init_collections():
    collections = [Keywords.T_CLINICALTRIALS_OBS.value, Keywords.T_CLINICALTRIALS_RAND.value,
                   Keywords.T_PUBLICATION_OBS.value, Keywords.T_PUBLICATION_RAND.value]
    for collection in collections:
        smc.get_db().drop_collection(collection)
        smc.get_db().create_collection(collection)


def load_excel_file():
    file_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(
        file_dir, 'excels', '20200601_IRIT_clinicalTrials+publications.xlsx')
    workbook = load_workbook(filename=file_path,)
    return workbook


def dump_all_excel_in_mongo():
    sheets = [Keywords.WS_CLINICALTRIALS_OBS.value, Keywords.WS_CLINICALTRIALS_RAND.value,
              Keywords.WS_PUBLICATION_OBS.value, Keywords.WS_PUBLICATION_RAND.value]
    collections = [Keywords.T_CLINICALTRIALS_OBS.value, Keywords.T_CLINICALTRIALS_RAND.value,
                   Keywords.T_PUBLICATION_OBS.value, Keywords.T_PUBLICATION_RAND.value]
    workbook = load_excel_file()
    for i, sheet in enumerate(sheets):
        dic_list = create_dic_by_sheet(sheet_object=workbook[sheet])
        smc.get_db()[collections[i]].insert_many(dic_list)


def test_insert_one_dic_in_mongo():
    dic = create_dic_by_sheet(sheet_object=load_excel_file()[
                              Keywords.WS_CLINICALTRIALS_OBS.value])[0]
    insert_one_dic_in_mongo(
        dic=dic, collection=Keywords.T_CLINICALTRIALS_OBS.value)


def insert_one_dic_in_mongo(*, dic, collection):
    smc.get_db()[collection].insert_one(dic)


def __main__():
    init_collections()
    dump_all_excel_in_mongo()


__main__()
