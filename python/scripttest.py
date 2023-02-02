from variables import Keywords
import pymongo as pm
from openpyxl import load_workbook, Workbook
import os
from urllib.parse import quote_plus


# file = pd.read_excel(
#     './others/20200601_IRIT_clinicalTrials+publications.xlsx', engine='openpyxl')


def iterating_over_values(path, sheet_name):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return

    sheet = workbook[sheet_name]
    for value in sheet.iter_rows(
            min_col=0, max_col=11, min_row=1, max_row=402,
            values_only=False):
        l = []
        for v in value:
            l.append(v.value)
        print(l)


def connection():
    username = quote_plus('<Jules>')
    password = quote_plus('<C6IyTitjpiYeq4t4>')
    cluster = '<SAE>'
    uri = 'mongodb+srv://' + username + ':' + password + '@' + cluster + \
        '.x1hujrr.mongodb.net'
    client = pm.MongoClient(uri)
    return client.test


def __main__():
    file_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(
        file_dir, 'excels', '20200601_IRIT_clinicalTrials+publications.xlsx')
    iterating_over_values(
        file_path, Keywords.WS_CLINICALTRIALS_OBS.value)
    # print(db)


__main__()
